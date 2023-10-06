"""
coding:utf-8
file: material.py
@time: 2023/9/17 23:36
@desc:
"""
import json
import os

from flask import Blueprint, request, url_for, escape, send_from_directory
from wms.decorators import get_params, path_existed
from wms.utils import ResultJson, config, get_uuid, const
from wms.models import MaterialSpec, User, Material, MaterialIn, Warehouse, MaterialOut, OperateLog
from urllib.parse import urljoin, unquote
from flask_jwt_extended import current_user, jwt_required
from openpyxl import load_workbook
from wms.plugins import db
from sqlalchemy.sql.expression import or_

material_bp = Blueprint('material_bp', __name__, url_prefix='/material')


@material_bp.route('/add/spec', methods=['POST'])
@path_existed(config.get('img.save'))
@jwt_required()
def add_spec():
    name = request.json.get('name')
    description = request.json.get('specification')
    uuid = request.json.get('uuid')
    images_path = os.path.join(config.get('img.save'), uuid)
    images = []
    if os.path.exists(images_path):
        images = os.listdir(images_path)
        images = [unquote(url_for('tools_bp.load_img', path=os.path.join(uuid, image))) for image in images]
    if MaterialSpec.query.filter_by(name=name).first():
        return ResultJson.bad_request(msg='规格名称已存在！')
    spec = MaterialSpec(
        name=name,
        description=description,
        images=images,
        user=current_user.id
    )
    spec.save()
    return ResultJson.ok(msg='规格添加成功！')


@material_bp.route('/list/spec')
@get_params(
    params=['page', 'size', 'name'],
    types=[int, int, str]
)
def list_spec(page, size, name):
    query = MaterialSpec.id
    if name:
        query = MaterialSpec.name.like(f"{name}%")
    specs = MaterialSpec.query.join(
        User,
        User.id == MaterialSpec.user
    ).filter(
        query
    ).with_entities(
        MaterialSpec.id,
        MaterialSpec.name,
        MaterialSpec.description,
        MaterialSpec.create_time,
        MaterialSpec.update_time,
        MaterialSpec.images,
        User.name.label('user')
    ).paginate(page=page, per_page=size)
    return ResultJson.ok(
        data=[dict(
            id=spec.id,
            name=spec.name,
            description=spec.description,
            images=[urljoin(config.HOST, img) for img in spec.images],
            create_time=str(spec.create_time),
            update_time=str(spec.update_time),
            user=spec.user
        ) for spec in specs.items],
        total=specs.total
    )


@material_bp.route('/stocking', methods=['POST'])
@get_params(
    params=['name', 'spec', 'unit', 'amount', 'price', 'warehouseId', 'type', 'code'],
    types=[str, str, str, str, str, str, str, str],
    methods='POST',
    check_para=True
)
@jwt_required()
def stocking_material(name, spec, unit, amount, price, warehouseId, type, code):
    material = Material(
        name=name,
        spec=spec,
        unit=unit,
        total=amount,
        left=amount,
        price=price,
        warehouse_id=warehouseId,
        type=type,
        barcode=code,
        user_id=current_user.id
    )
    material.save()
    MaterialIn(
        material_id=material.id,
        num=amount,
        user_id=current_user.id,
        warehouse_id=warehouseId
    ).save()
    return ResultJson.ok(msg='入库成功！')


@material_bp.route('/list')
@get_params(
    params=['page', 'size', 'name', 'spec', 'warehouse_id'],
    types=[int, int, str, str, str]
)
@jwt_required()
def material_list(page, size, name, spec, warehouse_id):
    query = (Material.id >= 0,)
    if name:
        query += (Material.name.like(f"{name}%"),)
    if spec:
        query += (Material.spec == spec,)
    if warehouse_id:
        query += (Material.warehouse_id == warehouse_id,)
    materials = Material.query.join(
        User,
        User.id == Material.user_id
    ).join(
        Warehouse,
        Warehouse.id == Material.warehouse_id
    ).join(
        MaterialSpec,
        MaterialSpec.id == Material.spec
    ).filter(*query).with_entities(
        Material.id,
        Material.name,
        MaterialSpec.name.label('spec'),
        Material.unit,
        Material.total,
        Material.left,
        Material.used,
        Material.price,
        Material.barcode,
        Material.type,
        Material.create_time,
        Material.update_time,
        User.name.label('user'),
        Warehouse.name.label('warehouse'),
        Material.status
    ).paginate(page=page, per_page=size)
    result = []
    for material in materials.items:
        result.append(dict(
            id=material.id,
            name=material.name,
            spec=material.spec,
            unit=material.unit,
            total=material.total,
            left=material.left,
            price=material.price,
            barcode=material.barcode,
            type=material.type,
            create_time=str(material.create_time),
            update_time=str(material.update_time),
            user=material.user,
            warehouse=material.warehouse,
            used=material.used,
            status=const.MATERIAL_STATUS.get(material.status)
        ))

    return ResultJson.ok(
        data=result,
        total=materials.total
    )


@material_bp.route('/get/<int:material_id>')
@jwt_required()
def material_detail(material_id):
    material = Material.query.join(
        User,
        User.id == Material.user_id
    ).join(
        Warehouse,
        Warehouse.id == Material.warehouse_id
    ).join(
        MaterialSpec,
        MaterialSpec.id == Material.spec
    ).filter(
        Material.id == material_id
    ).with_entities(
        Material.id,
        Material.name,
        MaterialSpec.name.label('spec'),
        MaterialSpec.description,
        User.name.label('user'),
        MaterialSpec.images,
        Material.unit,
        Material.total,
        Material.left,
        Material.used,
        Material.price,
        Material.barcode,
        Material.type,
        Material.create_time,
        Material.update_time,
        Warehouse.name.label('warehouse'),
        Warehouse.address,
        Warehouse.create_time.label('wc_time'),
        Warehouse.volume,
        Warehouse.status.label('w_status'),
        Warehouse.tag,
        Material.status
    ).first()
    return ResultJson.ok(
        data=dict(
            material=dict(
                id=material.id,
                name=material.name,
                type=material.type,
                price=material.price,
                barcode=material.barcode,
                unit=material.unit,
                total=material.total,
                left=material.left,
                used=material.used,
                create_time=str(material.create_time),
                update_time=str(material.update_time),
                user=material.user,
                status=const.MATERIAL_STATUS.get(material.status)
            ),
            spec=dict(
                name=material.spec,
                description=material.description,
                images=[urljoin(config.HOST, img) for img in material.images]
            ),
            warehouse=dict(
                name=material.warehouse,
                address=material.address,
                create_time=str(material.wc_time),
                volume=material.volume,
                tags=[tag for tag in material.tag],
                status=material.w_status
            )
        )
    )


@material_bp.route('/out', methods=['POST'])
@get_params(
    params=['material_id', 'out', 'reason'],
    types=[int, int, str],
    methods='POST',
    check_para=True
)
@jwt_required()
def out_material(material_id, out, reason):
    material = Material.query.filter_by(id=material_id).first()
    if not material:
        return ResultJson.not_found(msg='物料不存在！')
    if material.left < out:
        return ResultJson.bad_request(msg='物料数量不足！')
    material.left -= out
    material.used += out
    get_material_status(material)
    material.save()
    MaterialOut(
        material_id=material_id,
        num=out,
        reason=reason,
        user_id=current_user.id,
        warehouse_id=material.warehouse_id
    ).save()
    return ResultJson.ok(msg='出库成功！')


def get_material_status(material):
    if material.left / material.total > config.get('material.status.warning'):
        material.status = 0
    if material.left / material.total < config.get('material.status.warning'):
        material.status = 1
    if material.left / material.total < config.get('material.status.lack'):
        material.status = 2
    if material.left == 0:
        material.status = 3


@material_bp.route('/owner/out')
@get_params(
    params=['page', 'size', 'name', 'warehouse_id'],
    types=[int, int, str, str]
)
@jwt_required()
def owner_out_material(page, size, name, warehouse_id):
    query = (MaterialOut.user_id == current_user.id,)
    if name:
        query += (Material.name.like(f"{name}%"),)
    if warehouse_id:
        query += (Material.warehouse_id == warehouse_id,)
    materials = MaterialOut.query.join(
        User,
        User.id == MaterialOut.user_id
    ).join(
        Warehouse,
        Warehouse.id == MaterialOut.warehouse_id
    ).join(
        Material,
        Material.id == MaterialOut.material_id
    ).filter(*query).with_entities(
        MaterialOut.id,
        MaterialOut.num,
        MaterialOut.reason,
        MaterialOut.create_time,
        User.name.label('user'),
        Warehouse.name.label('warehouse'),
        Material.name.label('material')
    ).paginate(page=page, per_page=size)
    result = []
    for material in materials.items:
        result.append(dict(
            id=material.id,
            num=material.num,
            reason=material.reason,
            create_time=str(material.create_time),
            user=material.user,
            warehouse=material.warehouse,
            name=material.material
        ))
    return ResultJson.ok(
        data=result,
        total=materials.total
    )


@material_bp.route('/add/storage', methods=['POST'])
@jwt_required()
def add_storage():
    num = request.json.get('num')
    mid = request.json.get('mid')
    reason = request.json.get('reason')
    material = Material.query.filter(Material.id == mid).first()
    if not material:
        return ResultJson.not_found(msg='不存在的物料！')
    material.total += num
    material.left += num
    # material.save()
    get_material_status(material)
    material.save()
    OperateLog(
        user_id=current_user.id,
        type='add-storage',
        description=f'<p>给物料<b>{material.name}</b>添加了<b>{num}</b>库存。</p>',
        note=reason or ''
    ).save()
    return ResultJson.ok(msg='添加库存成功！')


@material_bp.route('/batch/template')
@path_existed(config.get('material.template'))
@jwt_required()
def batch_download():
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    warehouses = Warehouse.query.filter(Warehouse.status == 1).all()
    specs = MaterialSpec.query.all()
    workbook = Workbook()
    worksheet = workbook.active
    warehouse_options = [warehouse.name for warehouse in warehouses]
    spec_options = [spec.name for spec in specs]
    headers = ['物料名称', '物料类型', '物料编码', '物料单位', '物料单价', '入库数量', '所在仓库', '物料规格', '已用数量']
    worksheet.append(headers)

    def add_validation(options, worksheet, cell, placeholder='请选择'):
        validation = DataValidation(type='list', formula1=f'"{",".join(options)}"', allow_blank=True)
        worksheet.add_data_validation(validation)
        validation.add(cell)
        worksheet[cell] = placeholder

    add_validation(warehouse_options, worksheet, 'G2')
    add_validation(spec_options, worksheet, 'H2')
    worksheet['A2'] = '授权书纸张'
    worksheet['B2'] = '打印纸'
    worksheet['C2'] = 'SN20231541542'
    worksheet['D2'] = '张'
    worksheet['E2'] = 1.5
    worksheet['F2'] = 2000
    worksheet['I2'] = 0
    filename = get_uuid() + '.xlsx'
    workbook.save(os.path.join(config.get('material.template'), filename))
    return send_from_directory(
        directory=config.get('material.template'),
        path=filename,
        as_attachment=True
    )


@material_bp.route('/batch/stocking', methods=['POST'])
@path_existed(config.get('material.batch'))
@jwt_required()
def batch_stocking():
    if 'file' not in request.files:
        return ResultJson.bad_request(msg='请选择文件上传')
    file = request.files['file']
    same = request.form.get('same')
    if file.filename == '':
        return ResultJson.bad_request(msg='上传文件文件名不能为空')
    suffix = os.path.splitext(file.filename)[-1]
    if suffix != '.xlsx':
        return ResultJson.bad_request(msg='只支持xlsx格式的文件')
    filename = os.path.join(config.get('material.batch'), f'{get_uuid()}{suffix}')
    file.save(filename)
    headers = ['物料名称', '物料类型', '物料编码', '物料单位', '物料单价', '入库数量', '所在仓库', '物料规格', '已用数量']
    workbook = load_workbook(filename=filename)
    worksheet = workbook.active
    # 检查表头
    for index, cell in enumerate(worksheet[1]):
        if cell.value != headers[index]:
            return ResultJson.bad_request(msg='表头不正确，请下载模板后填写')
    materials = []
    # 读取数据
    for row in worksheet.iter_rows(min_row=2, max_col=9):
        material = {}
        for index, cell in enumerate(row):
            material[headers[index]] = cell.value
        materials.append(material)
    # 检查数据
    for idx, material in enumerate(materials):
        if not material.get('物料名称'):
            return ResultJson.bad_request(msg=f'第{idx+1}行，第{headers.index("物料名称")+1}列，物料名称不能为空')
        if not material.get('物料类型'):
            return ResultJson.bad_request(msg=f'第{idx+1}行，第{headers.index("物料类型")+1}列，物料类型不能为空')
        if not material.get('物料编码'):
            return ResultJson.bad_request(msg=f'第{idx+1}行，第{headers.index("物料编码")+1}列，物料编码不能为空')
        if not material.get('物料单位'):
            return ResultJson.bad_request(msg=f'第{idx+1}行，第{headers.index("物料单位")+1}列，物料单位不能为空')
        if not material.get('物料单价'):
            return ResultJson.bad_request(msg=f'第{idx+1}行，第{headers.index("物料单价")+1}列，物料单价不能为空')
        if not material.get('入库数量'):
            return ResultJson.bad_request(msg=f'第{idx+1}行，第{headers.index("入库数量")+1}列，入库数量不能为空')
        if not material.get('所在仓库'):
            return ResultJson.bad_request(msg=f'第{idx+1}行，第{headers.index("所在仓库")+1}列，所在仓库不能为空')
        if not material.get('物料规格'):
            return ResultJson.bad_request(msg=f'第{idx+1}行，第{headers.index("物料规格")+1}列，物料规格不能为空')
        if not material.get('已用数量') and material.get('已用数量') != 0:
            return ResultJson.bad_request(msg=f'第{idx+1}行，第{headers.index("已用数量")+1}列，已用数量不能为空')
        if not isinstance(material.get('入库数量'), int):
            return ResultJson.bad_request(msg=f'第{idx+1}行，第{headers.index("入库数量")+1}列，入库数量必须为整数')
        if not isinstance(material.get('已用数量'), int):
            return ResultJson.bad_request(msg=f'第{idx+1}行，第{headers.index("已用数量")+1}列，已用数量必须为整数')
        if material.get('所在仓库') not in [warehouse.name for warehouse in Warehouse.query.filter(Warehouse.status == 1).all()]:
            return ResultJson.bad_request(msg=f'第{idx+1}行，第{headers.index("所在仓库")+1}列，所在仓库错误')
        if material.get('物料规格') not in [spec.name for spec in MaterialSpec.query.all()]:
            return ResultJson.bad_request(msg=f'第{idx+1}行，第{headers.index("物料规格")+1}列，物料规格错误')
        if Material.query.filter(Material.barcode == material.get('物料编码')).first() and same == 'false':
            return ResultJson.bad_request(msg=f'第{idx+1}行，第{headers.index("物料编码")+1}列，物料编码已存在')
    print('开始写入数据库')
    # 写入数据库
    for material in materials:
        if same == 'true':
            name = material.get('物料名称')
            barcode = material.get('物料编码')
            db_material = Material.query.filter(or_(Material.barcode == barcode, Material.name == name)).first()
            if db_material:
                db_material.total += material.get('入库数量')
                db_material.used += material.get('已用数量')
                db_material.left += material.get('入库数量') - material.get('已用数量')
                db.session.commit()
                OperateLog(
                    user_id=current_user.id,
                    type='batch-stocking',
                    description=f'批量导入物料，物料名称：{material.get("物料名称")}，物料编码：{material.get("物料编码")}，入库数量：{material.get("入库数量")}，已用数量：{material.get("已用数量")}，所在仓库：{material.get("所在仓库")}，物料规格：{material.get("物料规格")}',
                    note='批量导入/编辑物料库存情况'
                ).save()
            else:
                insert_material(material)
        else:
            insert_material(material)
    return ResultJson.ok(msg='批量导入物料成功！')


def insert_material(material):
    db_warehouse = Warehouse.query.filter(Warehouse.name == material.get('所在仓库')).first()
    db_spec = MaterialSpec.query.filter(MaterialSpec.name == material.get('物料规格')).first()
    db_material = Material(
        name=material.get('物料名称'),
        type=material.get('物料类型'),
        barcode=material.get('物料编码'),
        unit=material.get('物料单位'),
        price=material.get('物料单价'),
        total=material.get('入库数量'),
        used=material.get('已用数量'),
        warehouse_id=db_warehouse.id,
        spec=db_spec.id,
        user_id=current_user.id,
        left=material.get('入库数量') - material.get('已用数量')
    )
    db_material.save()
    MaterialIn(
        material_id=db_material.id,
        warehouse_id=db_warehouse.id,
        num=material.get('入库数量'),
        user_id=current_user.id
    ).save()
